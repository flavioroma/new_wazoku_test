import pytest

from central.models import Challenge, Idea, IdeaVote, Site, User
from django.core import mail
from django.test import TestCase
from scripts import user_activity_export
from openpyxl import Workbook

class UserActivityExportTests(TestCase):

    def setUp(self):
        self.site, _ = Site.objects.get_or_create(domain="example.com")

        # Create some users
        self.contributor = User(
            username="contributor",
            email="contributor@example.com",
            first_name="Regular",
            last_name="User",
            is_contributor=True,
            site=self.site
        )
        self.contributor.save()

        self.manager = User(
            username="manager",
            email="manager@example.com",
            first_name="Mr",
            last_name="Manager",
            is_manager=True,
            site=self.site
        )
        self.manager.save()

        # Create a challenge and idea
        self.challenge = Challenge(
            name="Example challenge",
            description="A simple challenge for our tests",
            creator=self.manager,
            site=self.site
        )
        self.challenge.save()

        self.idea = Idea(
            name='Test idea for our challenge',
            summary='Simple idea',
            challenge=self.challenge,
            creator=self.contributor,
            site=self.site
        )
        self.idea.save()

        IdeaVote(creator=self.contributor, idea=self.idea).save()

    def test_export_script_sends_email(self):
        user_activity_export.main("example.com", "manager@example.com")

        # Test that one message has been sent.
        self.assertEqual(len(mail.outbox), 1)

        # Verify that the subject of the first message is correct.
        self.assertEqual(mail.outbox[0].subject, 'User Activity Export')

        # Test we have a single recipient
        self.assertEqual(len(mail.outbox[0].to), 1)

        # Verify that the recipient is the manager
        self.assertEqual(mail.outbox[0].to[0], "manager@example.com")

    def test_export_only_sends_emails_to_managers(self):
        self.manager.is_manager = False
        self.manager.save()
        user_activity_export.main("example.com", "manager@example.com")

        # Test that no message has been sent.
        self.assertEqual(len(mail.outbox), 0)

    def test_user_activity(self):
        wb = Workbook()
        ws = wb.active

        # create sheet 1 for active users
        ws1 = wb.create_sheet('Active users', 0)
        ws1['A1'] = 'Email'
        ws1['B1'] = 'Activity'
        ws1['A2'] = self.contributor.email
        ws1['B2'] = self.idea.name
        ws1['A3'] = self.manager.email
        ws1['B3'] = self.challenge.name

        # create sheet 2 for innactive users
        ws2 = wb.create_sheet('Innactive users', 1)
        ws2['A1'] = 'Users with no activity'
        wb.save("user_activity_export.xlsx")

    def test_multiple_manager(self):
        # add new user after first manager account
        manager_2 = User(
            username="manager2",
            email="manager2@example.com",
            first_name="Mr",
            last_name="Manager2",
            is_manager=True,
            site=self.site
        )
        manager_2.save()

        user_activity_export.main("example.com", "manager@example.com")

        self.assertEqual(len(mail.outbox), 1)

        self.assertEqual(mail.outbox[0].to[0], "manager@example.com")

    @pytest.mark.querytest
    def test_query_count(self):
        with self.assertNumQueries(12):
            user_activity_export.main("example.com", "manager@example.com")
